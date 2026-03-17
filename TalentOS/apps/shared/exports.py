"""CSV and Excel export utilities for ATS."""
import csv
import io
import logging
from datetime import datetime
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework import permissions
from drf_spectacular.utils import extend_schema

logger = logging.getLogger(__name__)


def _get_filename(base: str, fmt: str) -> str:
    ts = timezone.now().strftime('%Y%m%d_%H%M%S')
    return f'{base}_{ts}.{fmt}'


class ExportCandidatesView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=['Exports'], summary='Export candidates to CSV or Excel')
    def get(self, request):
        fmt = request.query_params.get('format', 'csv').lower()
        from apps.candidates.models import Candidate
        qs = Candidate.objects.filter(
            tenant=request.user.tenant
        ).select_related('tenant').order_by('-created_at')

        headers = ['ID', 'Full Name', 'Email', 'Phone', 'Source', 'Status', 'Created At']

        def row(c):
            return [
                str(c.id),
                c.full_name or '',
                c.email or '',
                getattr(c, 'primary_phone', '') or getattr(c, 'phone', '') or '',
                c.source or '',
                c.status or '',
                c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
            ]

        if fmt == 'xlsx':
            return _export_xlsx(qs, headers, row, 'candidates')
        return _export_csv(qs, headers, row, 'candidates')


class ExportJobsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=['Exports'], summary='Export jobs to CSV or Excel')
    def get(self, request):
        fmt = request.query_params.get('format', 'csv').lower()
        from apps.jobs.models import Job
        qs = Job.objects.filter(
            tenant=request.user.tenant
        ).order_by('-created_at')

        headers = ['ID', 'Title', 'Department', 'Location', 'Type', 'Status', 'Created At']

        def row(j):
            return [
                str(j.id),
                j.title or '',
                getattr(j, 'department', '') or '',
                getattr(j, 'location', '') or '',
                getattr(j, 'employment_type', '') or '',
                j.status or '',
                j.created_at.strftime('%Y-%m-%d %H:%M') if j.created_at else '',
            ]

        if fmt == 'xlsx':
            return _export_xlsx(qs, headers, row, 'jobs')
        return _export_csv(qs, headers, row, 'jobs')


class ExportApplicationsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=['Exports'], summary='Export applications to CSV or Excel')
    def get(self, request):
        fmt = request.query_params.get('format', 'csv').lower()
        from apps.applications.models import Application
        qs = Application.objects.filter(
            job__tenant=request.user.tenant
        ).select_related('candidate', 'job', 'current_stage').order_by('-created_at')

        headers = ['ID', 'Candidate', 'Email', 'Job Title', 'Stage', 'Status', 'Applied At']

        def row(a):
            return [
                str(a.id),
                a.candidate.full_name if a.candidate else '',
                a.candidate.email if a.candidate else '',
                a.job.title if a.job else '',
                a.current_stage.name if a.current_stage else '',
                a.status or '',
                a.created_at.strftime('%Y-%m-%d %H:%M') if a.created_at else '',
            ]

        if fmt == 'xlsx':
            return _export_xlsx(qs, headers, row, 'applications')
        return _export_csv(qs, headers, row, 'applications')


def _export_csv(qs, headers, row_fn, filename_base):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{_get_filename(filename_base, "csv")}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for obj in qs.iterator(chunk_size=500):
        writer.writerow(row_fn(obj))
    return response


def _export_xlsx(qs, headers, row_fn, filename_base):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return _export_csv(qs, headers, row_fn, filename_base)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename_base.capitalize()

    # Header row styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, obj in enumerate(qs.iterator(chunk_size=500), 2):
        for col_idx, value in enumerate(row_fn(obj), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{_get_filename(filename_base, "xlsx")}"'
    return response
