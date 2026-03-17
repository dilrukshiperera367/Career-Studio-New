"""CSV and Excel export utilities for HRM."""
import csv
import io
import logging
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework import permissions
from drf_spectacular.utils import extend_schema

logger = logging.getLogger(__name__)


def _get_filename(base: str, fmt: str) -> str:
    ts = timezone.now().strftime('%Y%m%d_%H%M%S')
    return f'{base}_{ts}.{fmt}'


class ExportEmployeesView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=['Exports'], summary='Export employees to CSV or Excel')
    def get(self, request):
        fmt = request.query_params.get('format', 'csv').lower()
        from core_hr.models import Employee
        qs = Employee.objects.filter(
            tenant=request.user.tenant
        ).select_related('department', 'position').order_by('last_name', 'first_name')

        headers = ['ID', 'Employee ID', 'First Name', 'Last Name', 'Email',
                   'Department', 'Position', 'Status', 'Hire Date']

        def row(e):
            return [
                str(e.id),
                getattr(e, 'employee_id', '') or '',
                e.first_name or '',
                e.last_name or '',
                e.email or '',
                e.department.name if getattr(e, 'department', None) else '',
                e.position.title if getattr(e, 'position', None) else '',
                getattr(e, 'status', '') or getattr(e, 'employment_status', '') or '',
                e.hire_date.strftime('%Y-%m-%d') if getattr(e, 'hire_date', None) else '',
            ]

        if fmt == 'xlsx':
            return _export_xlsx(qs, headers, row, 'employees')
        return _export_csv(qs, headers, row, 'employees')


class ExportLeaveRequestsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=['Exports'], summary='Export leave requests to CSV or Excel')
    def get(self, request):
        fmt = request.query_params.get('format', 'csv').lower()
        try:
            from leave_attendance.models import LeaveRequest
            qs = LeaveRequest.objects.filter(
                tenant=request.user.tenant
            ).select_related('employee').order_by('-created_at')
        except Exception:
            return HttpResponse('Export not available', status=503)

        headers = ['ID', 'Employee', 'Type', 'Start Date', 'End Date', 'Status', 'Requested At']

        def row(lr):
            emp = lr.employee if hasattr(lr, 'employee') else None
            return [
                str(lr.id),
                f'{emp.first_name} {emp.last_name}' if emp else '',
                getattr(lr, 'leave_type', '') or '',
                lr.start_date.strftime('%Y-%m-%d') if getattr(lr, 'start_date', None) else '',
                lr.end_date.strftime('%Y-%m-%d') if getattr(lr, 'end_date', None) else '',
                getattr(lr, 'status', '') or '',
                lr.created_at.strftime('%Y-%m-%d %H:%M') if lr.created_at else '',
            ]

        if fmt == 'xlsx':
            return _export_xlsx(qs, headers, row, 'leave_requests')
        return _export_csv(qs, headers, row, 'leave_requests')


class ExportPayrollView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=['Exports'], summary='Export payroll entries to CSV or Excel')
    def get(self, request):
        fmt = request.query_params.get('format', 'csv').lower()
        try:
            from payroll.models import PayrollEntry
            qs = PayrollEntry.objects.filter(
                payroll_run__tenant=request.user.tenant
            ).select_related('employee', 'payroll_run').order_by('-payroll_run__period_year',
                                                                   '-payroll_run__period_month')
        except Exception:
            return HttpResponse('Export not available', status=503)

        headers = ['ID', 'Employee', 'Period', 'Gross Pay', 'Net Pay', 'Status']

        def row(pe):
            emp = pe.employee if hasattr(pe, 'employee') else None
            run = pe.payroll_run if hasattr(pe, 'payroll_run') else None
            period = f'{run.period_year}-{run.period_month:02d}' if run else ''
            return [
                str(pe.id),
                f'{emp.first_name} {emp.last_name}' if emp else '',
                period,
                str(getattr(pe, 'gross_pay', '') or ''),
                str(getattr(pe, 'net_pay', '') or ''),
                getattr(pe, 'status', '') or '',
            ]

        if fmt == 'xlsx':
            return _export_xlsx(qs, headers, row, 'payroll')
        return _export_csv(qs, headers, row, 'payroll')


def _export_csv(qs, headers, row_fn, filename_base):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{_get_filename(filename_base, "csv")}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for obj in qs.iterator(chunk_size=500):
        writer.writerow(row_fn(obj))
    return response


def _export_xlsx(qs, headers, row_fn, filename_base):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return _export_csv(qs, headers, row_fn, filename_base)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename_base.capitalize()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, obj in enumerate(qs.iterator(chunk_size=500), 2):
        for col_idx, value in enumerate(row_fn(obj), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{_get_filename(filename_base, "xlsx")}"'
    return response
